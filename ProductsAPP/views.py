from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext as _
from django.contrib import messages
from django.conf import settings
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.views.decorators.csrf import csrf_exempt
from django.template.response import TemplateResponse
from django.contrib.contenttypes.models import ContentType

from io import BytesIO

import os
import json
from .models import BillAccount, Product, ProductFamily, BillPosition, Consumible, Refund, DiscountVoucher
from .forms import paymentMethodsForm, StockFormSet, ProductFormSet, barcode2BillForm, billSearchForm
from .tasks import printBillReceipt

import datetime
from UsersAPP.forms import findCustomerForm
from UsersAPP.models import Customer
from utils.usbUtils import ThermalPrinter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection


@login_required(login_url="login")
def add_bill(request):
    bill=BillAccount.create(createdBy=request.user)
    printer = ThermalPrinter()
    try:
        messages.info(request, printer.paperStatus)
    except:
        messages.warning(request, _("Cannot communicate with the receipt printer"))
    del printer
    return redirect('MaterialsAPP_edit_bill',code=bill.code,family_id=0)

@login_required(login_url="login")
def edit_bill(request,code,family_id=None,billPos=None):
    bill=BillAccount.objects.get(code=code)

    if not ProductFamily.objects.first():
        return TemplateResponse(request, 'errorPage.html',{'heading':_('Error on stock revision'),
                                                 'info':_('There are no product families created. Need to create at least one')
                                        })
    
    if not family_id:
        family_id = ProductFamily.objects.first().id

    productfamilies_tabs=[]
    productFamilies= ProductFamily.objects.all()
    for i,_type in enumerate(productFamilies):
        productfamilies_tabs.append({'name':_type.name,'id':_type.id,'active':_type.id==family_id})

    
    rows=Product.objects.filter(family=ProductFamily.objects.get(id=family_id))

    if billPos is None or billPos == 'None':
        billPos=None
    else:
        billPos=BillPosition.objects.get(id=int(billPos))


    return TemplateResponse(request, 'bill.html',{'bill' : bill,
                                        #'categories':non_emptyFamilies,
                                        'productfamilies_tabs':productfamilies_tabs,
                                        'rows':rows,
                                        'legend':"Categorías",
                                        "findCustomerForm":findCustomerForm(),
                                        "barcode2BillForm":barcode2BillForm(),
                                        "billPos":billPos
                                        })


@login_required(login_url="login")
def assign_customer(request,code):
    bill=BillAccount.objects.get(code=code)

    if request.method == 'POST':
        form = findCustomerForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data['data']
            customer = Customer.find(data=data)
            if customer:
                bill.setOwner(customer)
            else:
                messages.info(request, _("No customer found "))
                    
    return redirect('MaterialsAPP_edit_bill',code=bill.code,family_id=0)

@login_required(login_url="login")
def append_barcode_to_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    if request.method == 'POST':
        form = barcode2BillForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                product=Product.objects.get(barcode=barcode)
                return redirect('MaterialsAPP_append_to_bill',code=bill.code,id=product.id)
            except:
                pass            
    messages.error(request, _("The barcode introduced is not registered"))
    return redirect('MaterialsAPP_edit_bill',code = bill.code,tab=0) 



@login_required(login_url="login")
def append_to_bill(request,code,id):
    bill=BillAccount.objects.get(code=code)
    product=Product.objects.get(id=id)
    billPos=bill.add_bill_position(product=product,quantity=1)

    return redirect('MaterialsAPP_edit_billPos',code=bill.code,family_id=product.family.id,billPos=billPos.id)

@login_required(login_url="login")
def reduce_bill_position(request,id):
    bill_pos=BillPosition.objects.get(id=id)
    family=bill_pos.product.family
    bill_pos.reduce_quantity(quantity=1)
    return redirect('MaterialsAPP_edit_bill',code=bill_pos.bill.code,family_id=family.id)

@login_required(login_url="login")
def resume_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    paymentForm = paymentMethodsForm(instance=bill)
    return TemplateResponse(request, 'bill_resume.html',{'bill' : bill,'paymentForm':paymentForm,'vouchers':DiscountVoucher.objects.all()})

@login_required(login_url="login")
def print_bill(request,code):
    from myTPV.models import SiteSettings
    SETTINGS = SiteSettings.load()
    bill=BillAccount.objects.get(code=code)
    billData = bill.toJSON()
    from utils.pdfConverter import PrintedBill
    bill = PrintedBill(billData=billData,commerceData=SETTINGS.commerceData())
    filename = billData["code"]+".pdf"
    with open(filename, "wb") as binary_file:
        binary_file.write(bill.pdf)
    response = HttpResponse(open(filename, "rb"),content_type='application/force-download')
    response['Content-Disposition'] = 'attachment; filename=%s' % os.path.basename(filename)
    os.remove(billData["code"]+".pdf")
    return response

@login_required(login_url="login")
def print_receipt(request,code):
    bill=BillAccount.objects.get(code=code)
    billData = bill.toJSON()
    printBillReceipt.delay(billData=billData)
    return redirect('MaterialsAPP_resume_bill',code = bill.code)
    
@login_required(login_url="login")
def close_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    if request.method == 'POST':
        if bill.bill_positions.all().count() > 0:
            paymentForm = paymentMethodsForm(request.POST,instance=bill)
            if paymentForm.is_valid():
                paymentForm.save()
                bill.close()          
            else:
                messages.error(request, _("The payment method should be defined"))
                return redirect('MaterialsAPP_resume_bill',code = bill.code)
        else:
            bill.delete()

    return redirect('home')

@login_required(login_url="login")
def refund_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    if request.method == 'POST':
        if bill.bill_positions.all().count() > 0:
            paymentForm = paymentMethodsForm(request.POST,instance=bill)
            if paymentForm.is_valid():
                paymentForm.save()
                bill.close()          
            else:
                messages.error(request, _("The payment method should be defined"))
                return redirect('MaterialsAPP_resume_bill',code = bill.code)
    else:
        return TemplateResponse(request, 'bill_refund.html',{'bill' : bill,
                                        "barcode2BillForm":barcode2BillForm(),
                                        })

@login_required(login_url="login")
def refund_bill_position(request,id):
    bill_pos=BillPosition.objects.get(id=id)
    refund = Refund.objects.create(bill_pos=bill_pos)
    
    if refund.quantity<bill_pos.quantity:
        #refund.increaseQuantity(amount=1)
        messages.info(request, _("The product ")+str(bill_pos.product) + _(" has been refunded"))
    else:
        messages.error(request, _("All the items of the product ") + str(bill_pos.product) + _(" have been refunded"))
        
    return redirect('MaterialsAPP_refund_bill',code=bill_pos.bill.code)

@login_required(login_url="login")
def remove_barcode_to_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    if request.method == 'POST':
        form = barcode2BillForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            try:
                product=Product.objects.get(barcode=barcode)
                bill_pos = BillPosition.objects.get(product=product,bill=bill)
                return redirect('MaterialsAPP_refund_bill_position',id=bill_pos.id)
            except:
                pass            
    messages.error(request, _("The barcode introduced is not in the bill"))
    return redirect('MaterialsAPP_refund_bill',code = bill.code) 

@login_required(login_url="login")
def resume_refund(request,code):
    from myTPV.models import SiteSettings
    SETTINGS = SiteSettings.load()
    bill=BillAccount.objects.get(code=code)
    if SETTINGS.PRINT_RECEIPT_ON_REFUND:
        # here a printReceipt action could be issued
        billData = bill.toJSON()
        printBillReceipt.delay(billData=billData)
    Refund.close(bill=bill,user=request.user)
    return redirect('home')
    
@login_required(login_url="login")
def delete_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    bill.delete()
    return redirect('home')

@login_required(login_url="login")
def discountCredit_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    bill.discountUserCredit()
    messages.info(request, _("The customer's credit has been discounted"))
    return redirect('MaterialsAPP_resume_bill',code = bill.code)

@login_required(login_url="login")
def applyVoucher_bill(request,code,voucher_id):
    bill=BillAccount.objects.get(code=code)
    bill.applyVoucher(voucher_id)
    messages.info(request, _("The voucher has been discounted"))
    return redirect('MaterialsAPP_resume_bill',code = bill.code)

@login_required(login_url="login")
def cancelVoucher_bill(request,code):
    bill=BillAccount.objects.get(code=code)
    bill.cancelVouchers()
    messages.info(request, _("The vouchers have been cancelled"))
    return redirect('MaterialsAPP_resume_bill',code = bill.code)


def setMultiplier_billPosition(request,id):
    ''' This is an AJAX request, return data for the request to proceed '''
    if request.method == 'POST':
        value = json.loads(request.body.decode())['value']
        billPosition = BillPosition.objects.get(id=id)
        billPosition.set_quantity(quantity=value)
        response={'url':reverse('MaterialsAPP_edit_bill',kwargs={'code':billPosition.bill.code,'family_id':billPosition.product.family.id})}
    else:
        response=None
    return HttpResponse(json.dumps(response), content_type='application/json')

# Create your views here.
@login_required(login_url="login")
def check_stock(request,family_id=None):
    user = request.user
    
    if not ProductFamily.objects.first():
        return TemplateResponse(request, 'errorPage.html',{'heading':_('Error on stock revision'),
                                                 'info':_('There are no product families created. Need to create at least one')
                                        })
    
    if family_id is None:
        family_id = ProductFamily.objects.first().id

    productfamilies_tabs=[]
    productFamilies= ProductFamily.objects.all()
    for i,_type in enumerate(productFamilies):
        productfamilies_tabs.append({'name':_type.name,'id':_type.id,'active':_type.id==family_id})

    queryset=Consumible.objects.filter(infinite=False,family=ProductFamily.objects.get(id=family_id)).order_by('name')

    if request.method == 'POST':
        stock_formset = StockFormSet(request.POST,queryset=queryset)
        if stock_formset.is_valid():
            stock_formset.save()
            messages.success(request, _("The stock has been updated"))
            return redirect('home')
        else:
            for error in list(stock_formset.errors.values()):
                messages.error(request, error)

    stock_formset = StockFormSet(queryset=queryset)
    stock_value = Consumible.get_stock_value()
    return TemplateResponse(request, 'stock.html', {'productfamilies_tabs':productfamilies_tabs,
                                            'stock_formset': stock_formset,
                                          'stock_value':stock_value})

@login_required(login_url="login")
def check_products(request,family_id=None):
    user = request.user

    if not ProductFamily.objects.first():
        return TemplateResponse(request, 'errorPage.html',{'heading':_('Error on prices revision'),
                                                 'info':_('There are no product families created. Need to create at least one')
                                        })
    
    if family_id is None:
        family_id = ProductFamily.objects.first().id

    productfamilies_tabs=[]
    productFamilies= ProductFamily.objects.all()
    for i,_type in enumerate(productFamilies):
        productfamilies_tabs.append({'name':_type.name,'id':_type.id,'active':_type.id==family_id})

    queryset=Product.objects.filter(family=ProductFamily.objects.get(id=family_id))

    if request.method == 'POST':
        product_formset = ProductFormSet(request.POST,queryset=queryset)
        if product_formset.is_valid():
            product_formset.save()
            messages.success(request, _("Product prices have been updated"))
            return redirect('home')
        else:
            for error in list(product_formset.errors.values()):
                messages.error(request, error)

    product_formset = ProductFormSet(queryset=queryset)
    return TemplateResponse(request, 'products.html', {'productfamilies_tabs':productfamilies_tabs,
                                            'product_formset': product_formset,})

@login_required(login_url="login")
def historics_home(request):
    customer = None
    if request.method == "POST":
        form=billSearchForm(request.POST)
        payments = {}
        if form.is_valid():
            info={}
            info['code'] = form.cleaned_data['code']
            info['customer'] = form.cleaned_data['customer']
            info['to'] = form.cleaned_data['_to'] if form.cleaned_data['_to'] else datetime.datetime.today().date()+datetime.timedelta(days=1)
            info['from'] = form.cleaned_data['_from'] if form.cleaned_data['_from'] else info['to']-datetime.timedelta(days=365)
            if info['customer']:
                from myTPV.settings import get_customer_model
                try:
                    customer = get_customer_model().objects.get(email = info['customer'])
                except get_customer_model().DoesNotExist:
                    try:
                        customer = get_customer_model().objects.get(cif = info['customer'])
                    except get_customer_model().DoesNotExist:
                        customer = None
                if customer:
                    return redirect('UsersAPP_view_customer',id=customer.id)
                else:
                    bills = None
                    bill_totals=None
                    messages.error(request,_("No customer found"))
            elif info['code']:
                bills = BillAccount.objects.filter(code=info['code']).annotate(order_positions = Count('positions'))
                bill_totals=None
            else:
                bills = BillAccount.objects.filter(createdOn__gt=info['from'],
                                                   createdOn__lt=info['to']).annotate(order_positions = Count('positions'))
                
                for payment in BillAccount.PAYMENT_TYPES:
                    payments[payment[0]]={'value':0,'description':payment[1]}
        
                total=0
                total_vat=0
                for bill in bills.filter(status = BillAccount.STATUS_PAID):
                    total += bill.total
                    total_vat += bill.getVATAmount(withRefunds=False)
                    payments[bill.paymenttype]['value']=round(payments[bill.paymenttype]['value']+bill.total,2) 
                #bill_totals = bills.aggregate(total=Sum('total'),total_vat=Sum('vat_amount'))
                bill_totals={'total':total,'total_vat':total_vat}
            
            return TemplateResponse(request, 'historicBills.html', {'bills':bills,'number':bills.count() if bills else 0,
                                                          'totals':bill_totals,'payments':payments,'customer':customer,
                                                          'query_info':{'code':info['code'],'from':info['from'].isoformat(),'to':info['to'].isoformat()}})
            
    else:
        form=billSearchForm()

    return TemplateResponse(request, 'form.html', {'form': form,
                                        'title':_("Historic"),
                                        'back_to':'home',})

@csrf_exempt
def historics_download(request):
    if request.method == "POST":
        info = json.loads(request.body.decode())
        if info['code']:
            bills = BillAccount.objects.filter(code=info['code'])
        else:
            bills = BillAccount.objects.filter(createdOn__gt=info['from'],
                                                createdOn__lt=info['to'])
        
        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Bills', index=1)
        columns = ['Code', 'Date', 'Total', 'VAT', 'Customer', 'Payment']
        row_num = 1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        # Iterate through all coins
        for bill in bills:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                bill.code,
                bill.createdOn.replace(tzinfo=None),
                bill.total,
                bill.getVATAmount(),
                bill.owner.cif if bill.owner else "",
                bill.get_paymenttype_display(),
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(excelfile)
        response = HttpResponse(excelfile.getvalue(),content_type='application/force-download')
        filename = "facturas_"+str(info['from']).replace("-","_")+"_"+str(info['to']).replace("-","_")+".xlsx"
        response['Content-Disposition'] = 'attachment; filename='+filename
        return response